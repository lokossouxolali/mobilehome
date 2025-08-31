from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.utils.html import strip_tags
from django.shortcuts import redirect, render
from django.views import View
from .models import *
from django.contrib import messages
from django.db import transaction, models
import pdfkit
from django.template.loader import get_template
from .utils import pagination, get_invoice
import datetime
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from .decorators import *
from django.shortcuts import render, get_object_or_404, redirect
from decimal import Decimal
from django.contrib.auth.models import User
from django import forms
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Create your views here.
class HomeView(LoginRequiredSuperuserMixin, View):
    templates_name = 'index.html'
    invoices = Invoice.objects.select_related('customer', 'save_by').all().order_by('-invoice_date_time')
    
    def get(self, request, *args, **kwargs):
        # Calculer les statistiques
        total_invoices = Invoice.objects.count()
        total_customers = Customer.objects.count()
        total_articles = Article.objects.count()
        
        # Calculer le revenu total en utilisant les InvoiceItems
        total_revenue = InvoiceItem.objects.filter(
            invoice__paid=True
        ).aggregate(
            total=models.Sum(models.F('quantity') * models.F('unit_price'))
        )['total'] or 0
        
        items = pagination(request, self.invoices)
        
        context = {
            'invoices': items,
            'total_invoices': total_invoices,
            'total_customers': total_customers,
            'total_articles': total_articles,
            'total_revenue': total_revenue,
        }
        
        return render(request, self.templates_name, context)
    
    def post(self, request, *args, **kwargs):

        #Modify
        if request.POST.get('id_modified'):
            paid = request.POST.get('modified')

            try:
                obj = Invoice.objects.get(id=request.POST.get('id_modified'))
                if paid == 'True':
                    obj.paid = True
                else:
                    obj.paid = False
                obj.save()
                messages.success(request, "Modification effectuée avec succès.")

            except Exception as e:
                messages.error(request, f"Désolé, l'erreur suivante s'est produite : {e}.")


        # Suppression
        if request.POST.get('id_supprimer'):
            try:
                invoice_id = request.POST.get('id_supprimer')
                print(f"ID reçu pour suppression : {invoice_id}")  # Debugging
                obj = Invoice.objects.get(pk=invoice_id)
                obj.delete()
                messages.success(request, "La suppression a réussi.")
            except Invoice.DoesNotExist:
                messages.error(request, "La facture n'existe pas.")
            except Exception as e:
                messages.error(request, f"Désolé, l'erreur suivante s'est produite : {e}.")

        # Recalculer les statistiques après modification
        total_invoices = Invoice.objects.count()
        total_customers = Customer.objects.count()
        total_articles = Article.objects.count()
        
        # Calculer le revenu total en utilisant les InvoiceItems
        total_revenue = InvoiceItem.objects.filter(
            invoice__paid=True
        ).aggregate(
            total=models.Sum(models.F('quantity') * models.F('unit_price'))
        )['total'] or 0
        
        items = pagination(request, self.invoices)
        
        context = {
            'invoices': items,
            'total_invoices': total_invoices,
            'total_customers': total_customers,
            'total_articles': total_articles,
            'total_revenue': total_revenue,
        }
        
        return render(request, self.templates_name, context)
    

class AddCustomerView(LoginRequiredSuperuserMixin, View):
    template_name = 'add_customer.html'
    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)
    def post(self, request, *args, **kwargs):
        data = {
            'name': request.POST.get('name'),
            'email': request.POST.get('email'),
            'phone': request.POST .get('phone'),
            'address': request.POST.get('address'),
            'sex': request.POST.get('sex'),
            'age': request.POST.get('age'),
            'zip_code': request.POST.get('zip'),
            'save_by': request.user
        }
        try:
            created = Customer.objects.create(**data)
            if created:
                messages.success(request, "Le client s'est enregistré avec succès.")
                return redirect('add-invoice')
            else:
                messages.error(request, "Désolé, veuillez réessayer, les données envoyées sont corrompues.")
                return render(request, self.template_name)
        except Exception as e:
            messages.error(request, f"Désolé, le système détecte les problèmes suivants {e}.")
            return render(request, self.template_name)


from django.core.mail import EmailMessage
from django.utils.html import strip_tags
from django.templatetags.static import static
from django.core.mail import EmailMessage
from django.utils.html import strip_tags

class AddInvoiceView(LoginRequiredMixin, View):
    template_name = "add_invoice.html"

    def get(self, request, *args, **kwargs):
        customers = Customer.objects.all()
        articles = Article.objects.filter(is_active=True)
        context = {
            'customers': customers,
            'articles': articles
        }
        return render(request, self.template_name, context)

    @transaction.atomic()
    def post(self, request, *args, **kwargs):
        try:
            customer_id = request.POST.get('customer')
            invoice_type = request.POST.get('invoice_type')
            article_ids = request.POST.getlist('article')
            quantities = request.POST.getlist('qty')
            unit_prices = request.POST.getlist('unit')
            comments = request.POST.get('comment')

            customer = Customer.objects.get(id=customer_id)

            errors = []  # Liste pour stocker les erreurs

            # Vérification du stock AVANT de créer la facture
            for article_id, qty in zip(article_ids, quantities):
                article = Article.objects.get(id=article_id)
                qty = Decimal(qty)

                if article.stock < qty:
                    errors.append(f"Stock insuffisant pour l'article {article.name}. Stock disponible : {article.stock}.")

            if errors:
                for error in errors:
                    messages.error(request, error)
                    return redirect('add-invoice')
                return redirect('add-invoice')  # Redirection et arrêt du code si erreur

            # Si pas d'erreur, création de la facture
            invoice = Invoice.objects.create(
                customer=customer,
                save_by=request.user,
                invoice_type=invoice_type,
                comments=comments
            )

            invoice_items_details = [] #Creation d'une liste pour stocker les details de chaque article
            total_amount = 0 #initialisation du montant total
            domaine_name = request.build_absolute_uri('http://127.0.0.1:8000') #recuperation du nom de domaine
            # Création des InvoiceItems et mise à jour du stock
            for article_id, qty, unit_price in zip(article_ids, quantities, unit_prices):
                article = Article.objects.get(id=article_id)
                qty = Decimal(qty)
                unit_price = Decimal(unit_price)

                # Mise à jour du stock
                article.stock -= qty
                article.save()

                InvoiceItem.objects.create(
                    invoice=invoice,
                    article=article,
                    quantity=qty,
                    unit_price=unit_price
                )
                item_total = qty * unit_price #calcul du total pour chaque article
                total_amount += item_total #ajout du total de l'article au montant total
                invoice_items_details.append(f"""
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;">{article.name}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{qty}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{unit_price}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{item_total}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{article.stock}</td>
                    </tr>
                """) #Ajout des details de chaque article dans la liste

             # Construction du tableau HTML pour les détails des articles
            article_table = f"""
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background-color: #663399; color: white;">
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">ARTICLE</th>
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">QUANTITÉ</th>
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">PRIX UNITAIRE</th>
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">TOTAL</th>
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">STOCK RESTANT</th>
                        </tr>
                    </thead>
                    <tbody>
                        {''.join(invoice_items_details)}
                    </tbody>
                </table>
            """
            # Construction du corps du message HTML
            html_message = f"""
                <html>
                <head>
                    <style>
                        /* Ajoutez du CSS pour embellir l'e-mail si vous le souhaitez */
                        body {{ font-family: Arial, sans-serif; }}
                        .header {{
                            background-color: #663399;
                            color: white;
                            padding: 10px;
                            text-align: center; /* Centrer le texte */
                        }}
                        .content {{ padding: 10px; }}
                        .button {{
                            background-color: #4CAF50; /* Green */
                            border: none;
                            color: white;
                            padding: 10px 20px;
                            text-align: center;
                            text-decoration: none;
                            display: inline-block;
                            font-size: 16px;
                            margin: 4px 2px;
                            cursor: pointer;
                            border-radius: 5px;
                        }}
                        .footer {{
                            text-align: center; /* Centrer les boutons */
                            padding: 20px;
                        }}
                        .total-date {{
                            color: #6A0DAD;
                            font-size: 18px;
                            font-weight: bold;
                        }}
                    </style>
                </head>
                <body>
                    <div class="header">
                        <h2><span style="font-weight: bold; text-transform: uppercase;"> 📢 NOUVELLE FACTURE ENREGISTRÉE !</span></h2>
                    </div>
                    <div class="content">
                        <p>📢 Une nouvelle facture a été générée par <strong>{request.user.username}</strong> pour le client <strong>{customer.name}</strong>.</p>
                        <h3>Détails de la Facture :</h3>
                        {article_table}

                        <div class="total-date">
                            <p><strong>Montant Total :</strong> {total_amount}</p>
                            <p><strong>Date :</strong> {invoice.invoice_date_time}</p>
                        </div>

                        <!-- Bouton d'action -->
                        <div style="text-align: center; padding: 20px;  height: 100px;">
                            <a href="{domaine_name}/view_invoice/{invoice.id}" 
                            style="background-color: #27ae60; color: #fff; padding: 10px 20px; text-decoration: none; font-weight: bold; border-radius: 5px; display: inline-block; margin: 5px;">
                                📝 Voir la Facture
                            </a>
                            <a href="{domaine_name}/sales_summary_list" 
                            style="background-color: #27ae60; color: #fff; padding: 10px 20px; text-decoration: none; font-weight: bold; border-radius: 5px; display: inline-block; margin: 5px;">
                                📝 Voir les ventes
                            </a>
                            <a href="{domaine_name}/article_list" 
                            style="background-color: #27ae60; color: #fff; padding: 10px 20px; text-decoration: none; font-weight: bold; border-radius: 5px; display: inline-block; margin: 5px;">
                                📝 Voir les articles
                            </a>
                        </div>
                    </div>
                            <!-- Pied de page -->
                <div style="background-color: #6A0DAD; padding: 10px; text-align: center; font-size: 12px; color: #fff;">
                </body>
                </html>
            """

            plain_message = strip_tags(html_message)  # Version texte brut pour les clients de messagerie qui ne supportent pas HTML

            email = EmailMessage(
                subject="Nouvelle vente effectuée",
                body=html_message, #utilisation de html_message et non plain_message pour le corps de l'email
                from_email=request.user.email,
                to=["lokossouxolali@gmail.com"],
            )
            email.content_subtype = "html"  # Indique que le contenu est HTML
            email.send()

            header_image_url = static("images/mobilehouseacceuil.svg")

            # Icônes des réseaux sociaux (remplace les liens avec ceux de tes réseaux)
            social_media_icons = """
            <div style="text-align: center; padding: 10px;">
                <a href="https://facebook.com/tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/174/174848.png" width="24" height="24">
                </a>
                <a href="https://instagram.com/tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/2111/2111463.png" width="24" height="24">
                </a>
                <a href="https://twitter.com/tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/733/733579.png" width="24" height="24">
                </a>
                <a href="https://linkedin.com/tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/174/174857.png" width="24" height="24">
                </a>
                <a href="https://www.tiktok.com/@tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/3046/3046125.png" width="24" height="24">
                </a>
                <a href="https://wa.me/+22896393780" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/733/733585.png" width="24" height="24">
                </a>
                <a href="https://www.youtube.com/c/tonpage" style="margin: 0 5px;">
                    <img src="https://cdn-icons-png.flaticon.com/24/1384/1384060.png" width="24" height="24">
                </a>
            </div>
            """

            # Corps du message HTML
            html_message = f"""
            <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 0;
                        padding: 0;
                    }}
                    body {{
                        font-family: 'Arial', sans-serif;
                        background-color: #1c1c1c; /* Noir clair ultra classe */
                        color: #f1f1f1;
                        text-align: center;
                        padding: 40px;
                    }}
                    h1 {{
                        color: #6A0DAD;
                    }}
                    p {{
                        font-size: 16px;
                        line-height: 1.6;
                    }}
                    .header {{
                        text-align: center;
                        background-color : #6A0DAD;
                        padding: 20px;
                    }}
                    .content {{
                        padding: 20px;
                        text-align: center;
                    }}
                    .button {{
                        display: inline-block;
                        background-color: #27ae60;
                        color: white;
                        padding: 10px 20px;
                        text-decoration: none;
                        font-weight: bold;
                        border-radius: 5px;
                        margin-top: 20px;
                    }}
                    .footer {{
                        text-align: center;
                        background-color: #f1f1f1;
                        padding: 20px;
                        font-size: 12px;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <img src="{header_image_url}" width="100%" alt="En-tête">
                </div>
                    <h1>🛍️ Merci pour votre achat, {customer.name} !</h1>
                    <p>Nous sommes ravis de vous compter parmi nos précieux clients. Votre confiance signifie tout pour nous !</p>
                    <p>🌟 Chaque achat chez nous est une promesse de qualité et de satisfaction.</p>
                    <p>Si vous avez la moindre question, nous sommes là pour vous.</p>

                {social_media_icons}

                <div class="footer">
                    <p>Besoin d'aide ? <a href="mailto:lokossouxolali@gmail.com">Contactez-nous</a></p>
                    <p>© 2025 Mobile House. Tous droits réservés.</p>
                </div>
            </body>
            </html>
            """

            # Version texte brut (au cas où l'email HTML ne s'affiche pas)
            plain_message = strip_tags(html_message)

            # Envoi de l'email
            email = EmailMessage(
                subject="Merci pour votre achat !",
                body=html_message,
                from_email="lokossouxolali@gmail.com",
                to=[customer.email],
            )
            email.content_subtype = "html"  # Indique que l'email est au format HTML
            email.send()


            messages.success(request, "Facture créée avec succès.")
            return redirect('home')

        except Exception as e:
            messages.error(request, f"Désolé, une erreur s'est produite : {e}.")
            return render(request, self.template_name)


class InvoiceVisualizationView(LoginRequiredSuperuserMixin, View):
    template_name = 'invoice.html'
    def get(self,request, *args, **kwargs):
        pk = kwargs.get('pk')
        context = get_invoice(pk)

        return render(request, self.template_name, context)


@superuser_required  
def get_invoice_pdf(request, *args, **kwargs):
        #Generer un fichier pdf a partir de html
        pk =kwargs.get('pk')
        context = get_invoice(pk)
        context['date'] = datetime.datetime.today()
        #Get html file
        template = get_template('invoice_pdf.html')
        #Render html with context variable : Envoyer les variables de context directement vers le template
        html = template.render(context)
        #Option of pdf format
        options = {
            'page-size':'Letter',
            'encoding':'UTF-8',
            'enable-local-file-access': ''
        }

        #Generate pdf
        try:
        # Générer le PDF
            # Configuration du chemin vers wkhtmltopdf
            config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
            pdf = pdfkit.from_string(html, False, options, configuration=config)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
            return response

        except Exception as e:
            return HttpResponse(f"Erreur lors de la génération du PDF : {e}", status=500)
        

# Formulaire personnalisé pour créer un administrateur
class AdminCreationForm(forms.ModelForm):
    class Meta:
        model = User
        fields = ['username', 'email', 'password']

    def save(self, commit=True):
        user = super().save(commit=False)
        user.is_staff = True  # Définit l'utilisateur comme administrateur
        user.set_password(self.cleaned_data['password'])  # Hache le mot de passe
        if commit:
            user.save()
        return user

# Vue pour ajouter un administrateur
@login_required
def add_admin(request):
    if request.method == 'POST':
        form = AdminCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Administrateur créé avec succès !")
            return redirect('admin_list')  # Remplacez par votre URL cible
    else:
        form = AdminCreationForm()

    return render(request, 'add_admin.html', {'form': form})

# views.py
@login_required
def admin_list(request):
    admins = User.objects.filter(is_staff=True)
    admins = pagination(request, admins)  # Paginer les admins
    return render(request, 'admin_list.html', {'admins': admins})

# Vue pour ajouter un nouvel article
def add_article(request):
    if request.method == "POST":
        try:
            # Récupération des données du formulaire
            name = request.POST.get('name')
            stock = request.POST.get('stock')

            # Création de l'article dans la base de données
            article = Article.objects.create(
                name=name,
                stock=stock
            )

            # Message de succès
            messages.success(request, "L'article a été ajouté avec succès.")
            return redirect('article-list')  # Rediriger vers la même page ou une autre page après l'ajout

        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout de l'article : {e}")

    # Si la requête est en GET, on affiche simplement le formulaire
    return render(request, "add_article.html")

def article_list(request):
    articles = Article.objects.filter(is_active=True)  # Récupérer tous les produits

    if request.method == 'POST' and 'id_supprimer' in request.POST:
        try:
            article_id = request.POST.get('id_supprimer')
            print(f"ID reçu pour suppression : {article_id}")  # Debugging
            # Get the Article object
            article = Article.objects.get(pk=article_id)
            # Delete the Article object
            article.is_active = False
            article.save()

                    # Construire l'e-mail HTML
            article_table = f"""
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background-color: #d9534f; color: white;">
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">ARTICLE SUPPRIMÉ</th>
                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">STOCK AVANT SUPPRESSION</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd;">{article.name}</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">{article.stock}</td>
                        </tr>
                    </tbody>
                </table>
            """

            html_message = f"""
                <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; }}
                        .header {{
                            background-color: #d9534f;
                            color: white;
                            padding: 10px;
                            text-align: center;
                        }}
                        .content {{ padding: 10px; }}
                    </style>
                </head>
                <body>
                    <div class="header">
                        <h2>🚨 ARTICLE SUPPRIMÉ</h2>
                    </div>
                    <div class="content">
                        <p><strong>{request.user.username}</strong> a supprimé un article :</p>
                        {article_table}
                        <p>❌ Cet article a été désactivé dans le système.</p>
                    </div>
                </body>
                </html>
            """

            plain_message = strip_tags(html_message)

            # Envoi de l'e-mail
            email = EmailMessage(
                subject="Un article a été supprimé",
                body=html_message,
                from_email=request.user.email,
                to=["lokossouxolali@gmail.com"],  # Remplace par l'email de l'administrateur
            )
            email.content_subtype = "html"
            email.send()

            messages.success(request, "L'article a été supprimé avec succès.")
        except Article.DoesNotExist:
            messages.error(request, "L'article n'existe pas.")
        except Exception as e:
            messages.error(request, f"Désolé, l'erreur suivante s'est produite : {e}.")


    articles = pagination(request, articles)  # Paginer les articles
    return render(request, 'article_list.html', {'articles': articles})

def customer_list(request):
    customers = Customer.objects.all()  # Récupérer tous les produits
    customers = pagination(request, customers)  # Paginer les clients
    return render(request, 'customer_list.html', {'customers': customers})

def sales_summary(request):
    # Paginer directement les InvoiceItems
    invoice_items = InvoiceItem.objects.all().order_by('-invoice__invoice_date_time')
    sales_data = pagination(request, invoice_items)
    
    context = {
        'sales_data': sales_data
    }
    return render(request, 'sales_summary_list.html', context)


@superuser_required
def export_sales_pdf(request):
    """Export des données de vente en PDF"""
    try:
        # Récupérer toutes les données de vente
        sales_data = InvoiceItem.objects.all().order_by('-invoice__invoice_date_time')
        
        # Calculer les statistiques
        total_sales = sales_data.count()
        total_revenue = sum(sale.total_price for sale in sales_data)
        total_quantity = sum(sale.quantity for sale in sales_data)
        unique_customers = sales_data.values('invoice__customer').distinct().count()
        
        # Préparer le contexte
        context = {
            'sales_data': sales_data,
            'total_sales': total_sales,
            'total_revenue': total_revenue,
            'total_quantity': total_quantity,
            'unique_customers': unique_customers,
            'date': datetime.datetime.now()
        }
        
        # Générer le PDF
        template = get_template('sales_pdf.html')
        html = template.render(context)
        
        # Options de format PDF
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in'
        }
        
        # Configuration du chemin vers wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf = pdfkit.from_string(html, False, options, configuration=config)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_ventes_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {e}")
        return redirect('sales-summary-list')


@superuser_required
def export_articles_pdf(request):
    """Export de la liste des articles en PDF"""
    try:
        # Récupérer tous les articles
        articles = Article.objects.filter(is_active=True)
        
        # Calculer les statistiques
        total_articles = articles.count()
        total_stock = sum(article.stock for article in articles)
        in_stock = articles.filter(stock__gt=0).count()
        out_of_stock = articles.filter(stock=0).count()
        
        # Préparer le contexte
        context = {
            'articles': articles,
            'total_articles': total_articles,
            'total_stock': total_stock,
            'in_stock': in_stock,
            'out_of_stock': out_of_stock,
            'date': datetime.datetime.now()
        }
        
        # Générer le PDF
        template = get_template('articles_pdf.html')
        html = template.render(context)
        
        # Options de format PDF
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in'
        }
        
        # Configuration du chemin vers wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf = pdfkit.from_string(html, False, options, configuration=config)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_articles_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {e}")
        return redirect('article-list')


@superuser_required
def export_customers_pdf(request):
    """Export de la liste des clients en PDF"""
    try:
        # Récupérer tous les clients
        customers = Customer.objects.all()
        
        # Calculer les statistiques
        total_customers = customers.count()
        active_customers = customers.filter(invoice__isnull=False).distinct().count()
        total_purchases = Invoice.objects.count()
        total_revenue = InvoiceItem.objects.filter(
            invoice__paid=True
        ).aggregate(
            total=models.Sum(models.F('quantity') * models.F('unit_price'))
        )['total'] or 0
        
        # Préparer le contexte
        context = {
            'customers': customers,
            'total_customers': total_customers,
            'active_customers': active_customers,
            'total_purchases': total_purchases,
            'total_revenue': total_revenue,
            'date': datetime.datetime.now()
        }
        
        # Générer le PDF
        template = get_template('customers_pdf.html')
        html = template.render(context)
        
        # Options de format PDF
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in'
        }
        
        # Configuration du chemin vers wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf = pdfkit.from_string(html, False, options, configuration=config)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_clients_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {e}")
        return redirect('customer-list')


@superuser_required
def export_admins_pdf(request):
    """Export de la liste des administrateurs en PDF"""
    try:
        # Récupérer tous les administrateurs
        admins = User.objects.filter(is_staff=True)
        
        # Calculer les statistiques
        total_admins = admins.count()
        active_admins = admins.filter(is_active=True).count()
        superusers = admins.filter(is_superuser=True).count()
        staff_users = admins.filter(is_staff=True, is_superuser=False).count()
        
        # Préparer le contexte
        context = {
            'admins': admins,
            'total_admins': total_admins,
            'active_admins': active_admins,
            'superusers': superusers,
            'staff_users': staff_users,
            'date': datetime.datetime.now()
        }
        
        # Générer le PDF
        template = get_template('admins_pdf.html')
        html = template.render(context)
        
        # Options de format PDF
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in'
        }
        
        # Configuration du chemin vers wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf = pdfkit.from_string(html, False, options, configuration=config)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_administrateurs_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {e}")
        return redirect('admin-list')


@superuser_required
def export_dashboard_pdf(request):
    """Export du tableau de bord en PDF"""
    try:
        # Récupérer les données du tableau de bord
        invoices = Invoice.objects.select_related('customer', 'save_by').all().order_by('-invoice_date_time')
        
        # Calculer les statistiques
        total_invoices = Invoice.objects.count()
        total_customers = Customer.objects.count()
        total_articles = Article.objects.count()
        total_revenue = InvoiceItem.objects.filter(
            invoice__paid=True
        ).aggregate(
            total=models.Sum(models.F('quantity') * models.F('unit_price'))
        )['total'] or 0
        
        # Préparer le contexte
        context = {
            'invoices': invoices,
            'total_invoices': total_invoices,
            'total_customers': total_customers,
            'total_articles': total_articles,
            'total_revenue': total_revenue,
            'date': datetime.datetime.now()
        }
        
        # Générer le PDF
        template = get_template('dashboard_pdf.html')
        html = template.render(context)
        
        # Options de format PDF
        options = {
            'page-size': 'A4',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'margin-top': '0.5in',
            'margin-right': '0.5in',
            'margin-bottom': '0.5in',
            'margin-left': '0.5in'
        }
        
        # Configuration du chemin vers wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf = pdfkit.from_string(html, False, options, configuration=config)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="tableau_bord_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {e}")
        return redirect('home')


def edit_article(request, article_id):
    article = get_object_or_404(Article, pk=article_id)
    old_stock = article.stock

    if request.method == 'POST':
        name = request.POST.get('name')
        stock = request.POST.get('stock')

        # Validation manuelle
        errors = {}
        if not name:
            errors['name'] = "Le nom de l'article est obligatoire."
        if not stock:
            errors['stock'] = "Le stock est obligatoire."
        try:
            stock = int(stock)
            if stock < 0:
                errors['stock'] = "Le stock doit être un nombre positif."
        except ValueError:
            errors['stock'] = "Le stock doit être un nombre entier."

        if errors:
            context = {'article': article, 'errors': errors, 'name': name, 'stock': stock}
            return render(request, 'edit_article.html', context)

        # Si la validation réussit, enregistrez les modifications
        article.name = name
        article.stock = stock
        article.save()

                # Envoi d'un e-mail à l'administrateur
        article_table = f"""
            <table style="width: 100%; border-collapse: collapse;">
                <thead>
                    <tr style="background-color: #663399; color: white;">
                        <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">ARTICLE</th>
                        <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">STOCK AVANT</th>
                        <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">STOCK APRÈS</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;">{article.name}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{old_stock}</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{stock}</td>
                    </tr>
                </tbody>
            </table>
        """

        html_message = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; }}
                    .header {{
                        background-color: #663399;
                        color: white;
                        padding: 10px;
                        text-align: center;
                    }}
                    .content {{ padding: 10px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h2>📢 MODIFICATION D'ARTICLE</h2>
                </div>
                <div class="content">
                    <p><strong>{request.user.username}</strong> a modifié un article :</p>
                    {article_table}
                    <p>✅ Les modifications ont été appliquées avec succès.</p>
                </div>
            </body>
            </html>
        """

        plain_message = strip_tags(html_message)  

        email = EmailMessage(
            subject="Article modifié",
            body=html_message,
            from_email=request.user.email,
            to=["lokossouxolali@gmail.com"],
        )
        email.content_subtype = "html"
        email.send()

        messages.success(request, "L'article a été modifié avec succès.")
        return redirect('article-list')  # Redirige vers la liste des articles
    else:
        # Affiche le formulaire pré-rempli
        context = {'article': article, 'name': article.name, 'stock': article.stock}
        return render(request, 'edit_article.html', context)


@superuser_required
def export_sales_excel(request):
    """Export des données de vente en format Excel"""
    try:
        # Récupérer toutes les données de vente
        sales_data = InvoiceItem.objects.select_related('invoice__customer', 'article').all().order_by('-invoice__invoice_date_time')
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventes"
        
        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Styles pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre du rapport
        ws.merge_cells('A1:G1')
        ws['A1'] = "RAPPORT DES VENTES - MOBILE HOUSE"
        ws['A1'].font = Font(bold=True, size=16, color="6366F1")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Date de génération
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=12, color="64748B")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes des colonnes
        headers = [
            "N°", "Article", "Quantité Vendue", "Date de Vente", 
            "Client", "Prix Unitaire (FCFA)", "Prix Total (FCFA)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données des ventes
        total_revenue = 0
        total_quantity = 0
        
        for row, sale in enumerate(sales_data, 5):
            # Numéro de ligne
            ws.cell(row=row, column=1, value=row-4).alignment = data_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            # Article
            ws.cell(row=row, column=2, value=sale.article.name).alignment = data_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            # Quantité
            ws.cell(row=row, column=3, value=sale.quantity).alignment = data_alignment
            ws.cell(row=row, column=3).border = thin_border
            total_quantity += sale.quantity
            
            # Date de vente
            sale_date = sale.invoice.invoice_date_time.strftime('%d/%m/%Y %H:%M')
            ws.cell(row=row, column=4, value=sale_date).alignment = data_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            # Client
            ws.cell(row=row, column=5, value=sale.invoice.customer.name).alignment = data_alignment
            ws.cell(row=row, column=5).border = thin_border
            
            # Prix unitaire
            ws.cell(row=row, column=6, value=float(sale.unit_price)).alignment = data_alignment
            ws.cell(row=row, column=6).border = thin_border
            
            # Prix total
            total_price = float(sale.total_price)
            ws.cell(row=row, column=7, value=total_price).alignment = data_alignment
            ws.cell(row=row, column=7).border = thin_border
            total_revenue += total_price
        
        # Ligne de total
        total_row = len(sales_data) + 5
        ws.cell(row=total_row, column=1, value="").border = thin_border
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3, value=total_quantity).font = Font(bold=True)
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=4, value="").border = thin_border
        ws.cell(row=total_row, column=5, value="").border = thin_border
        ws.cell(row=total_row, column=6, value="").border = thin_border
        ws.cell(row=total_row, column=7, value=total_revenue).font = Font(bold=True)
        ws.cell(row=total_row, column=7).border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 30, 15, 20, 25, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ventes_mobile_house_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder le fichier
        wb.save(response)
        
        messages.success(request, "Export Excel des ventes généré avec succès !")
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {e}")
        return redirect('sales-summary')


@superuser_required
def export_articles_excel(request):
    """Export des données d'articles en format Excel"""
    try:
        # Récupérer toutes les données d'articles
        articles_data = Article.objects.filter(is_active=True).all().order_by('-created_at')
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles"
        
        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Styles pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre du rapport
        ws.merge_cells('A1:E1')
        ws['A1'] = "RAPPORT DES ARTICLES - MOBILE HOUSE"
        ws['A1'].font = Font(bold=True, size=16, color="6366F1")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Date de génération
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=12, color="64748B")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes des colonnes
        headers = [
            "N°", "Nom de l'Article", "Stock", "Date d'Ajout", "Stock Vendus"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données des articles
        total_stock = 0
        total_sold = 0
        
        for row, article in enumerate(articles_data, 5):
            # Numéro de ligne
            ws.cell(row=row, column=1, value=row-4).alignment = data_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            # Nom de l'article
            ws.cell(row=row, column=2, value=article.name).alignment = data_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            # Stock
            ws.cell(row=row, column=3, value=article.stock).alignment = data_alignment
            ws.cell(row=row, column=3).border = thin_border
            total_stock += article.stock
            
            # Date d'ajout
            created_date = article.created_at.strftime('%d/%m/%Y')
            ws.cell(row=row, column=4, value=created_date).alignment = data_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            # Stock vendus
            ws.cell(row=row, column=5, value=article.total_sold).alignment = data_alignment
            ws.cell(row=row, column=5).border = thin_border
            total_sold += article.total_sold
        
        # Ligne de total
        total_row = len(articles_data) + 5
        ws.cell(row=total_row, column=1, value="").border = thin_border
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3, value=total_stock).font = Font(bold=True)
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=4, value="").border = thin_border
        ws.cell(row=total_row, column=5, value=total_sold).font = Font(bold=True)
        ws.cell(row=total_row, column=5).border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 35, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="articles_mobile_house_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder le fichier
        wb.save(response)
        
        messages.success(request, "Export Excel des articles généré avec succès !")
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {e}")
        return redirect('article-list')


@superuser_required
def export_customers_excel(request):
    """Export des données de clients en format Excel"""
    try:
        # Récupérer toutes les données de clients
        customers_data = Customer.objects.all().order_by('-id')
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Styles pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre du rapport
        ws.merge_cells('A1:F1')
        ws['A1'] = "RAPPORT DES CLIENTS - MOBILE HOUSE"
        ws['A1'].font = Font(bold=True, size=16, color="6366F1")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Date de génération
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=12, color="64748B")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes des colonnes
        headers = [
            "N°", "Nom du Client", "Email", "Téléphone", "Adresse", "Nombre d'Achats"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données des clients
        total_customers = 0
        total_purchases = 0
        
        for row, customer in enumerate(customers_data, 5):
            # Numéro de ligne
            ws.cell(row=row, column=1, value=row-4).alignment = data_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            # Nom du client
            ws.cell(row=row, column=2, value=customer.name).alignment = data_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            # Email
            ws.cell(row=row, column=3, value=customer.email or "").alignment = data_alignment
            ws.cell(row=row, column=3).border = thin_border
            
            # Téléphone
            ws.cell(row=row, column=4, value=customer.phone or "").alignment = data_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            # Adresse
            address = f"{customer.address}, {customer.city}" if customer.address and customer.city else (customer.address or customer.city or "")
            ws.cell(row=row, column=5, value=address).alignment = data_alignment
            ws.cell(row=row, column=5).border = thin_border
            
            # Nombre d'achats
            purchase_count = customer.invoice_set.count()
            ws.cell(row=row, column=6, value=purchase_count).alignment = data_alignment
            ws.cell(row=row, column=6).border = thin_border
            total_purchases += purchase_count
            
            total_customers += 1
        
        # Ligne de total
        total_row = len(customers_data) + 5
        ws.cell(row=total_row, column=1, value="").border = thin_border
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3, value="").border = thin_border
        ws.cell(row=total_row, column=4, value="").border = thin_border
        ws.cell(row=total_row, column=5, value="").border = thin_border
        ws.cell(row=total_row, column=6, value=total_purchases).font = Font(bold=True)
        ws.cell(row=total_row, column=6).border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 25, 25, 15, 30, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="clients_mobile_house_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder le fichier
        wb.save(response)
        
        messages.success(request, "Export Excel des clients généré avec succès !")
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {e}")
        return redirect('customer-list')


@superuser_required
def export_admins_excel(request):
    """Export des données d'administrateurs en format Excel"""
    try:
        # Récupérer toutes les données d'administrateurs
        admins_data = User.objects.filter(is_staff=True).all().order_by('-date_joined')
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Administrateurs"
        
        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Styles pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre du rapport
        ws.merge_cells('A1:E1')
        ws['A1'] = "RAPPORT DES ADMINISTRATEURS - MOBILE HOUSE"
        ws['A1'].font = Font(bold=True, size=16, color="6366F1")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Date de génération
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=12, color="64748B")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes des colonnes
        headers = [
            "N°", "Nom d'Utilisateur", "Email", "Date d'Inscription", "Statut"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données des administrateurs
        total_admins = 0
        superusers = 0
        staff_users = 0
        
        for row, admin in enumerate(admins_data, 5):
            # Numéro de ligne
            ws.cell(row=row, column=1, value=row-4).alignment = data_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            # Nom d'utilisateur
            ws.cell(row=row, column=2, value=admin.username).alignment = data_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            # Email
            ws.cell(row=row, column=3, value=admin.email or "").alignment = data_alignment
            ws.cell(row=row, column=3).border = thin_border
            
            # Date d'inscription
            join_date = admin.date_joined.strftime('%d/%m/%Y')
            ws.cell(row=row, column=4, value=join_date).alignment = data_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            # Statut
            if admin.is_superuser:
                status = "Super Admin"
                superusers += 1
            elif admin.is_staff:
                status = "Staff"
                staff_users += 1
            else:
                status = "Utilisateur"
            
            ws.cell(row=row, column=5, value=status).alignment = data_alignment
            ws.cell(row=row, column=5).border = thin_border
            
            total_admins += 1
        
        # Ligne de total
        total_row = len(admins_data) + 5
        ws.cell(row=total_row, column=1, value="").border = thin_border
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3, value="").border = thin_border
        ws.cell(row=total_row, column=4, value="").border = thin_border
        ws.cell(row=total_row, column=5, value=total_admins).font = Font(bold=True)
        ws.cell(row=total_row, column=5).border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 25, 30, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="admins_mobile_house_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder le fichier
        wb.save(response)
        
        messages.success(request, "Export Excel des administrateurs généré avec succès !")
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {e}")
        return redirect('admin-list')


@superuser_required
def export_dashboard_excel(request):
    """Export des données du tableau de bord en format Excel"""
    try:
        # Récupérer toutes les données
        invoices_data = Invoice.objects.select_related('customer').all().order_by('-invoice_date_time')[:10]  # 10 dernières factures
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Tableau de Bord"
        
        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Styles pour les bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre du rapport
        ws.merge_cells('A1:E1')
        ws['A1'] = "TABLEAU DE BORD - MOBILE HOUSE"
        ws['A1'].font = Font(bold=True, size=16, color="6366F1")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Date de génération
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=12, color="64748B")
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Statistiques générales
        ws.merge_cells('A4:E4')
        ws['A4'] = "STATISTIQUES GÉNÉRALES"
        ws['A4'].font = Font(bold=True, size=14, color="1E293B")
        ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Calculer les statistiques
        total_invoices = Invoice.objects.count()
        total_customers = Customer.objects.count()
        total_articles = Article.objects.count()
        total_revenue = InvoiceItem.objects.filter(
            invoice__paid=True
        ).aggregate(
            total=models.Sum(models.F('quantity') * models.F('unit_price'))
        )['total'] or 0
        
        # Statistiques
        stats_data = [
            ["Total des Factures", total_invoices],
            ["Total des Clients", total_customers],
            ["Total des Articles", total_articles],
            ["Chiffre d'Affaires (FCFA)", f"{total_revenue:.2f}"]
        ]
        
        for row, (label, value) in enumerate(stats_data, 5):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=value).border = thin_border
            ws.cell(row=row, column=3, value="").border = thin_border
            ws.cell(row=row, column=4, value="").border = thin_border
            ws.cell(row=row, column=5, value="").border = thin_border
        
        # Factures récentes
        ws.merge_cells('A9:E9')
        ws['A9'] = "FACTURES RÉCENTES"
        ws['A9'].font = Font(bold=True, size=14, color="1E293B")
        ws['A9'].alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes des colonnes pour les factures
        invoice_headers = [
            "N°", "Client", "Date", "Montant (FCFA)", "Statut"
        ]
        
        for col, header in enumerate(invoice_headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données des factures récentes
        for row, invoice in enumerate(invoices_data, 12):
            # Numéro de ligne
            ws.cell(row=row, column=1, value=row-11).alignment = data_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            # Client
            ws.cell(row=row, column=2, value=invoice.customer.name).alignment = data_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            # Date
            invoice_date = invoice.invoice_date_time.strftime('%d/%m/%Y %H:%M')
            ws.cell(row=row, column=3, value=invoice_date).alignment = data_alignment
            ws.cell(row=row, column=3).border = thin_border
            
            # Montant
            total_amount = invoice.total_amount
            ws.cell(row=row, column=4, value=f"{total_amount:.2f}").alignment = data_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            # Statut
            status = "Payée" if invoice.paid else "En attente"
            ws.cell(row=row, column=5, value=status).alignment = data_alignment
            ws.cell(row=row, column=5).border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 25, 20, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tableau_bord_mobile_house_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder le fichier
        wb.save(response)
        
        messages.success(request, "Export Excel du tableau de bord généré avec succès !")
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {e}")
        return redirect('home')